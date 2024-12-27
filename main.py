import re
import openpyxl
from openpyxl import Workbook
from tkinter import filedialog

def read_stock(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_start_row = 2
    data = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if row[0] is None:
            continue
        data.append({
            "仓库": row[0],
            "商品名称": row[1],
            "货号": row[2],
            "尺码": row[3],
            "成本价": row[4],
            "库存": row[5],
            "当前毒普通价": row[6],
            "价格更新时间": row[7],
            "3.5到手": row[8],
            "4.0到手": row[9],
            "5.0到手": row[10],
            "入库时间": row[11],
            "备注": row[12],
        })
    return data

def read_dewu(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = "销售订单"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在！有效的工作表为: {workbook.sheetnames}")
    sheet = workbook[sheet_name]

    data_start_row = 4
    data = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        if len(row) < 58 or row[3] is None or row[5] is None or row[57] is None:
            continue
        data.append({
            "商品货号": row[3],
            "数量": row[4],
            "规格": row[5],
            "实付金额": row[57],
        })
    return data

def get_normalized_size(size_str):
    special_sizes = {
        '⅓': '.3',
        '⅔': '.5',
    }
    
    # 先尝试提取数字部分
    size = re.search(r'\d+', str(size_str))
    if not size:
        return str(size_str)
    
    base_size = size.group()
    
    # 检查是否有特殊分数
    for fraction, decimal in special_sizes.items():
        if fraction in size_str:
            # 如果是 ⅓，向下取整（比如37⅓ -> 37）
            if fraction == '⅓':
                return base_size
            # 如果是 ⅔，加0.5（比如36⅔ -> 36.5）
            elif fraction == '⅔' or fraction == '½':
                return f"{base_size}.5"
    
    # 如果没有特殊分数，返回原始数字
    return base_size

def compare_and_calculate(data_stock, data_dewu):
    # 创建一个字典来存储得物订单，键为(货号, 规格)组合
    dewu_dict = {}
    for row_dewu in data_dewu:
        # 处理规格
        size = get_normalized_size(row_dewu["规格"])
        
        key = (row_dewu["商品货号"], size)
        if key not in dewu_dict:
            dewu_dict[key] = []
        dewu_dict[key].append(row_dewu)

    results = []
    for row_stock in data_stock:
        # 确保库存中的尺码也使用相同的标准化处理
        stock_size = get_normalized_size(row_stock["尺码"])
        key = (row_stock["货号"], stock_size)
        
        if key in dewu_dict:
            stock = int(row_stock["库存"]) if row_stock["库存"] is not None else 0
            remaining_stock = stock

            # 处理该商品的所有订单
            for row_dewu in dewu_dict[key]:
                cost_price = float(row_stock["成本价"]) if row_stock["成本价"] is not None else 0
                paid_amount = float(row_dewu["实付金额"]) if row_dewu["实付金额"] is not None else 0
                
                difference = round(paid_amount) - round(cost_price)
                sold_quantity = min(row_dewu["数量"], remaining_stock)
                remaining_stock -= sold_quantity

                result = row_stock.copy()
                result["利润"] = difference * sold_quantity
                result["库存"] = remaining_stock
                result["卖出数量"] = sold_quantity
                results.append(result)

                if remaining_stock <= 0:
                    break
        else:
            results.append(row_stock)

    return results

def write_to_excel(data, headers, output_path, include_profit):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "计算结果" if include_profit else "导入文件"

    # 写入表头
    if include_profit:
        sheet.append(headers + ["利润", "卖出数量"])
        # 只写入有利润数据的记录
        filtered_data = [row for row in data if "利润" in row and row["利润"] != 0]
    else:
        sheet.append(headers)
        filtered_data = data

    # 写入数据
    for row_data in filtered_data:
        row = [row_data.get(header, "") for header in headers]
        if include_profit:
            row.append(row_data.get("利润", ""))
            row.append(row_data.get("卖出数量", ""))
        sheet.append(row)

    workbook.save(output_path)

def main():
    # 文件选择对话框
    print("请选择库存文件...")
    stock_path = filedialog.askopenfilename(
        title="选择库存文件",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not stock_path:
        print("未选择库存文件，程序退出")
        return

    print("请选择得物订单文件...")
    dewu_path = filedialog.askopenfilename(
        title="选择得物订单文件",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not dewu_path:
        print("未选择得物订单文件，程序退出")
        return

    # 设置输出文件名（与脚本在同一目录）
    output_profit_path = "利润结果.xlsx"
    output_import_path = "导入文件.xlsx"

    try:
        data_stock = read_stock(stock_path)
        data_dewu = read_dewu(dewu_path)

        headers = ["仓库", "商品名称", "货号", "尺码", "成本价", "库存", 
                  "当前毒普通价", "价格更新时间", "3.5到手", "4.0到手", 
                  "5.0到手", "入库时间", "备注"]

        results = compare_and_calculate(data_stock, data_dewu)

        # 生成输出文件
        write_to_excel(results, headers, output_profit_path, include_profit=True)
        write_to_excel(results, headers, output_import_path, include_profit=False)

        print(f"\n处理完成！文件已生成：")
        print(f"- {output_profit_path}（包含利润和库存）")
        print(f"- {output_import_path}（库存已更新，无利润列）")
        
    except Exception as e:
        print(f"\n发生错误：{e}")

if __name__ == "__main__":
    main()
